"""
Espace dédié aux entreprises clientes.
Gestion de leurs agents, congés, permissions, présences et exports Excel.
"""
from ._base import *
from django.core.paginator import Paginator
from django.db.models import Count, Q as Qdb


@login_required
@rh_required
def entreprise_list(request):
    """Liste des entreprises avec nombre d'agents (admin uniquement).

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprises/list.html.
"""
    q = request.GET.get('q', '')
    manager_direction = get_manager_direction(request.user)

    if manager_direction:
        ent_ids = Contractuel.objects.filter(
            direction=manager_direction
        ).values_list('entreprise_id', flat=True).distinct()
        entreprises_qs = Entreprise.objects.filter(pk__in=ent_ids)
    else:
        entreprises_qs = Entreprise.objects.all()
    if q:
        entreprises_qs = entreprises_qs.filter(nom__icontains=q)

    # Annotation en 1 seule requête (évite N+1)
    if manager_direction:
        entreprises_qs = entreprises_qs.annotate(
            nb_actifs=Count('contractuels', filter=Qdb(
                contractuels__direction=manager_direction,
                contractuels__statut='ACTIF')),
            nb_total=Count('contractuels', filter=Qdb(
                contractuels__direction=manager_direction)),
        )
    else:
        entreprises_qs = entreprises_qs.annotate(
            nb_actifs=Count('contractuels', filter=Qdb(contractuels__statut='ACTIF')),
            nb_total=Count('contractuels'),
        )
    entreprises_data = [
        {'e': e, 'nb_actifs': e.nb_actifs, 'nb_total': e.nb_total}
        for e in entreprises_qs
    ]

    paginator = Paginator(entreprises_data, 12)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'SYGEPECO/entreprises/list.html', {
        'entreprises_data': page_obj, 'page_obj': page_obj,
        'q': q, 'manager_direction': manager_direction,
    })


@login_required
@rh_required
def entreprise_detail(request, pk):
    """Fiche détaillée d'une entreprise avec ses contractuels actifs.

    Args:
        request: HttpRequest Django.
        pk (int): Clé primaire de l'entreprise.

    Returns:
        HttpResponse : template entreprises/detail.html.
"""
    entreprise = get_object_or_404(Entreprise, pk=pk)
    contractuels = entreprise.contractuels.select_related('poste', 'direction').all()
    md = get_manager_direction(request.user)
    if md:
        contractuels = contractuels.filter(direction=md)
    return render(request, 'SYGEPECO/entreprises/detail.html', {
        'entreprise': entreprise, 'contractuels': contractuels,
    })


@login_required
@administrateur_required
def entreprise_create(request):
    """Crée une nouvelle entreprise (admin uniquement).

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : formulaire ou redirection.
"""
    from ..forms import EntrepriseForm
    form = EntrepriseForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        e = form.save()
        log_action(request.user, f"Creation entreprise : {e.nom}", 'Entreprise', e.pk)
        messages.success(request, f"Entreprise {e.nom} creee.")
        return redirect('entreprise_detail', pk=e.pk)
    return render(request, 'SYGEPECO/entreprises/form.html',
                  {'form': form, 'titre': 'Nouvelle entreprise'})


@login_required
@administrateur_required
def entreprise_update(request, pk):
    """Modifie une entreprise existante (admin uniquement).

    Args:
        request: HttpRequest Django.
        pk (int): Clé primaire de l'entreprise.

    Returns:
        HttpResponse : formulaire pré-rempli ou redirection.
"""
    from ..forms import EntrepriseForm
    entreprise = get_object_or_404(Entreprise, pk=pk)
    form = EntrepriseForm(request.POST or None, request.FILES or None, instance=entreprise)
    if request.method == 'POST' and form.is_valid():
        form.save()
        log_action(request.user, f"Modification entreprise : {entreprise.nom}", 'Entreprise', entreprise.pk)
        messages.success(request, "Entreprise mise a jour.")
        return redirect('entreprise_detail', pk=entreprise.pk)
    return render(request, 'SYGEPECO/entreprises/form.html', {
        'form': form, 'titre': f'Modifier - {entreprise.nom}', 'entreprise': entreprise})


@login_required
@administrateur_required
def entreprise_delete(request, pk):
    """Supprime une entreprise (admin uniquement, POST).

    Args:
        request: HttpRequest Django (POST).
        pk (int): Clé primaire de l'entreprise.

    Returns:
        HttpResponseRedirect vers la liste.
"""
    entreprise = get_object_or_404(Entreprise, pk=pk)
    if request.method == 'POST':
        nom = entreprise.nom
        entreprise.delete()
        log_action(request.user, f"Suppression entreprise : {nom}", 'Entreprise')
        messages.success(request, f"Entreprise {nom} supprimee.")
        return redirect('entreprise_list')
    return render(request, 'SYGEPECO/entreprises/confirm_delete.html', {'entreprise': entreprise})


@login_required
def entreprise_espace_home(request):
    """Tableau de bord de l'espace Entreprise.

    Affiche : nombre d'agents actifs, congés EN_ATTENTE,
    présences du jour, contrats expirant bientôt.

    Accès : @entreprise_required.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/home.html.
"""
    if not hasattr(request.user, 'profile') or request.user.profile.role not in (
            'ENTREPRISE', 'ADMINISTRATEUR', 'DRH'):
        messages.error(request, "Acces non autorise.")
        return redirect('login')

    entreprise = get_entreprise_for_user(request.user)
    entreprises = Entreprise.objects.filter(active=True)
    agents = Contractuel.objects.filter(entreprise=entreprise) if entreprise else Contractuel.objects.all()
    today = date.today()

    dept_stats = (agents.values('direction__nom').annotate(total=Count('id')).order_by('-total')[:5])
    alertes_conges = build_alertes_conges(Conge.objects.filter(contractuel__in=agents), today)

    return render(request, 'SYGEPECO/entreprise_espace/home.html', {
        'entreprise': entreprise, 'entreprises': entreprises,
        'total': agents.count(),
        'actifs': agents.filter(statut='ACTIF').count(),
        'conges_attente': Conge.objects.filter(contractuel__in=agents, statut='EN_ATTENTE').count(),
        'permissions_attente': Permission.objects.filter(contractuel__in=agents, statut='EN_ATTENTE').count(),
        'presences_today': Presence.objects.filter(contractuel__in=agents, date=today, statut='PRESENT').count(),
        'dept_stats': dept_stats, 'alertes_conges': alertes_conges,
    })


@login_required
def entreprise_espace_agents(request):
    """Liste des agents actifs de l'entreprise connectée.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/agents.html.
"""
    if not hasattr(request.user, 'profile') or request.user.profile.role not in (
            'ENTREPRISE', 'ADMINISTRATEUR', 'DRH'):
        messages.error(request, "Acces non autorise.")
        return redirect('login')
    entreprise = get_entreprise_for_user(request.user)
    agents = Contractuel.objects.select_related('poste', 'direction', 'entreprise')
    if entreprise:
        agents = agents.filter(entreprise=entreprise)
    q = request.GET.get('q', '')
    if q:
        agents = agents.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(matricule__icontains=q))
    statut = request.GET.get('statut', '')
    if statut:
        agents = agents.filter(statut=statut)
    return render(request, 'SYGEPECO/entreprise_espace/agents.html', {
        'entreprise': entreprise, 'agents': agents,
        'q': q, 'statut': statut, 'total': agents.count(),
    })


@login_required
def entreprise_espace_conges(request):
    """Gestion des congés des agents de l'entreprise.

    Filtrable par statut. Inclut les boutons Approuver/Rejeter.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/conges.html.
"""
    if not hasattr(request.user, 'profile') or request.user.profile.role not in (
            'ENTREPRISE', 'ADMINISTRATEUR', 'DRH'):
        messages.error(request, "Acces non autorise.")
        return redirect('login')
    entreprise = get_entreprise_for_user(request.user)
    conges = Conge.objects.select_related('contractuel', 'approuve_par')
    if entreprise:
        conges = conges.filter(contractuel__entreprise=entreprise)
    statut = request.GET.get('statut', '')
    if statut:
        conges = conges.filter(statut=statut)
    a_approuver = Conge.objects.filter(statut='VALIDE_MANAGER')
    en_attente  = Conge.objects.filter(statut='EN_ATTENTE')
    if entreprise:
        a_approuver = a_approuver.filter(contractuel__entreprise=entreprise)
        en_attente  = en_attente.filter(contractuel__entreprise=entreprise)
    return render(request, 'SYGEPECO/entreprise_espace/conges.html', {
        'entreprise': entreprise, 'conges': conges.order_by('-created_at'), 'statut': statut,
        'conges_a_approuver': a_approuver.count(), 'conges_attente': en_attente.count(),
    })


@login_required
def entreprise_espace_permissions(request):
    """Gestion des permissions des agents de l'entreprise.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/permissions.html.
"""
    if not hasattr(request.user, 'profile') or request.user.profile.role not in (
            'ENTREPRISE', 'ADMINISTRATEUR', 'DRH'):
        messages.error(request, "Acces non autorise.")
        return redirect('login')
    entreprise = get_entreprise_for_user(request.user)
    perms = Permission.objects.select_related('contractuel', 'approuve_par')
    if entreprise:
        perms = perms.filter(contractuel__entreprise=entreprise)
    statut = request.GET.get('statut', '')
    if statut:
        perms = perms.filter(statut=statut)
    en_attente = Permission.objects.filter(statut='EN_ATTENTE')
    if entreprise:
        en_attente = en_attente.filter(contractuel__entreprise=entreprise)
    return render(request, 'SYGEPECO/entreprise_espace/permissions.html', {
        'entreprise': entreprise, 'permissions': perms.order_by('-created_at'), 'statut': statut,
        'permissions_attente': en_attente.count(),
    })


@login_required
def entreprise_espace_presences(request):
    """Présences des agents de l'entreprise pour une date donnée.

    Paramètre GET : `date` (défaut : aujourd'hui).

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/presences.html.
"""
    if not hasattr(request.user, 'profile') or request.user.profile.role not in (
            'ENTREPRISE', 'ADMINISTRATEUR', 'DRH'):
        messages.error(request, "Acces non autorise.")
        return redirect('login')
    entreprise = get_entreprise_for_user(request.user)
    today = date.today()
    date_str = request.GET.get('date', str(today))
    try:
        date_filtre = date.fromisoformat(date_str)
    except ValueError:
        date_filtre = today
    presences = Presence.objects.select_related('contractuel').filter(date=date_filtre)
    if entreprise:
        presences = presences.filter(contractuel__entreprise=entreprise)
    return render(request, 'SYGEPECO/entreprise_espace/presences.html', {
        'entreprise': entreprise,
        'presences': presences.order_by('contractuel__nom'),
        'date_filtre': date_filtre,
    })


@login_required
def entreprise_espace_presence_create(request):
    """Enregistre une présence pour un agent de l'entreprise.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : formulaire ou redirection.
"""
    if not hasattr(request.user, 'profile') or request.user.profile.role not in (
            'ENTREPRISE', 'ADMINISTRATEUR', 'DRH'):
        messages.error(request, "Acces non autorise.")
        return redirect('login')
    entreprise = get_entreprise_for_user(request.user)
    date_retour = request.GET.get('date', str(date.today()))
    form = PresenceForm(request.POST or None)
    if entreprise:
        form.fields['contractuel'].queryset = Contractuel.objects.filter(
            entreprise=entreprise, statut='ACTIF').order_by('nom', 'prenom')
    if request.method == 'POST' and form.is_valid():
        presence = form.save()
        log_action(request.user,
                   f"Presence enregistree pour {presence.contractuel} ({presence.date})",
                   'Presence', presence.pk)
        messages.success(request, f"Presence enregistree pour {presence.contractuel}.")
        return redirect(f"{reverse('entreprise_espace_presences')}?date={presence.date}")
    return render(request, 'SYGEPECO/entreprise_espace/presence_form.html', {
        'entreprise': entreprise, 'form': form, 'date_retour': date_retour,
    })


@login_required
def entreprise_espace_contrats(request):
    """Contrats actifs des agents de l'entreprise.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/contrats.html.
"""
    entreprise, err = ent_check(request)
    if err: return err
    qs = Contrat.objects.select_related('contractuel', 'type_contrat').all()
    if entreprise:
        qs = qs.filter(contractuel__entreprise=entreprise)
    statut = request.GET.get('statut', '')
    if statut:
        qs = qs.filter(statut=statut)
    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(
            Q(contractuel__nom__icontains=q) | Q(contractuel__prenom__icontains=q) |
            Q(contractuel__matricule__icontains=q))
    return render(request, 'SYGEPECO/entreprise_espace/contrats.html', {
        'contrats': qs, 'entreprise': entreprise, 'statut': statut, 'q': q,
    })


@login_required
def entreprise_espace_rapports(request):
    """Page des rapports et exports de l'espace Entreprise.

    Args:
        request: HttpRequest Django.

    Returns:
        HttpResponse : template entreprise_espace/rapports.html.
"""
    entreprise, err = ent_check(request)
    if err: return err
    today = date.today()
    return render(request, 'SYGEPECO/entreprise_espace/rapports.html', {
        'entreprise': entreprise, 'mois': today.month, 'annee': today.year,
        'mois_courant': today.month, 'annee_courante': today.year,
    })


def _xlsx_export(request, qs_fn, headers, row_fn, col_widths, filename_prefix):
    """Helper générique pour les exports Excel de l'espace Entreprise.

    Args:
        request: HttpRequest Django.
        qs_fn (callable): Fonction retournant le QuerySet à exporter.
        headers (list[str]): En-têtes des colonnes.
        row_fn (callable): Fonction (obj) → liste de valeurs par ligne.
        col_widths (list[int]): Largeurs des colonnes.
        filename_prefix (str): Préfixe du nom de fichier Excel.

    Returns:
        HttpResponse : fichier .xlsx en attachment.
"""
    entreprise, err = ent_check(request)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installe.", status=500)

    qs = qs_fn(entreprise)
    wb = openpyxl.Workbook()
    ws = wb.active
    thin = Side(style='thin', color='333344')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(bold=True, color="D4A853", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.row_dimensions[1].height = 22
    for row, obj in enumerate(qs, 2):
        data, fc = row_fn(obj)
        rf = PatternFill("solid", fgColor=fc)
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border; cell.fill = rf
            cell.alignment = Alignment(
                horizontal='center' if col > 3 else 'left', vertical='center')
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ent_slug = entreprise.nom.replace(' ', '_') if entreprise else 'toutes'
    today = date.today()
    mois = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    fname = f"{filename_prefix}_{ent_slug}_{mois:02d}_{annee}.xlsx"
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


@login_required
def entreprise_export_presences(request):
    """Exporte les présences de l'entreprise en Excel.

    Paramètre GET : `date` (défaut : aujourd'hui).

    Args:
        request: HttpRequest Django.

    Returns:
        FileResponse : fichier Excel.
"""
    entreprise, err = ent_check(request)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installe.", status=500)
    today = date.today()
    mois  = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    qs = Presence.objects.select_related('contractuel').filter(date__month=mois, date__year=annee)
    if entreprise:
        qs = qs.filter(contractuel__entreprise=entreprise)
    qs = qs.order_by('contractuel__nom', 'date')
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Presences {mois:02d}-{annee}"
    thin = Side(style='thin', color='333344')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(bold=True, color="D4A853", size=11)
    headers = ['Matricule','Nom','Prenom','Date','Arrivee','Depart','Statut','Observations']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
    ws.row_dimensions[1].height = 22
    LABELS = {'PRESENT':'Present','ABSENT':'Absent','RETARD':'Retard',
              'CONGE':'Conge','PERMISSION':'Permission','JUSTIFIE':'Justifie'}
    for row, p in enumerate(qs, 2):
        data = [p.contractuel.matricule,p.contractuel.nom,p.contractuel.prenom,
                p.date.strftime('%d/%m/%Y'),
                str(p.heure_arrivee)[:5] if p.heure_arrivee else '-',
                str(p.heure_depart)[:5] if p.heure_depart else '-',
                LABELS.get(p.statut,p.statut), p.observations or '']
        fc = "F0FFF4" if p.statut=='PRESENT' else ("FFF0F0" if p.statut=='ABSENT' else "FFFBF0")
        rf = PatternFill("solid", fgColor=fc)
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border; cell.fill = rf
            cell.alignment = Alignment(horizontal='center' if col>3 else 'left', vertical='center')
    for i, ww in enumerate([14,20,20,12,10,10,14,30], 1):
        ws.column_dimensions[get_column_letter(i)].width = ww
    ent_slug = entreprise.nom.replace(' ','_') if entreprise else 'toutes'
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="presences_{ent_slug}_{mois:02d}_{annee}.xlsx"'
    wb.save(resp); return resp


@login_required
def entreprise_export_conges(request):
    """Exporte les congés de l'entreprise en Excel.

    Args:
        request: HttpRequest Django.

    Returns:
        FileResponse : fichier Excel.
"""
    entreprise, err = ent_check(request)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installe.", status=500)
    today = date.today()
    mois  = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    qs = Conge.objects.select_related('contractuel','approuve_par').filter(
        date_debut__month=mois, date_debut__year=annee)
    if entreprise:
        qs = qs.filter(contractuel__entreprise=entreprise)
    qs = qs.order_by('contractuel__nom')
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Conges {mois:02d}-{annee}"
    thin = Side(style='thin', color='333344')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(bold=True, color="D4A853", size=11)
    headers = ['Matricule','Nom','Prenom','Type','Debut','Fin','Jours','Statut','Approuve par']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
    ws.row_dimensions[1].height = 22
    STATUS = {'EN_ATTENTE':'En attente','APPROUVE':'Approuve','REJETE':'Rejete'}
    for row, c in enumerate(qs, 2):
        jours = (c.date_fin-c.date_debut).days+1 if c.date_fin and c.date_debut else ''
        approuve = (c.approuve_par.get_full_name() or c.approuve_par.username
                    if c.approuve_par else '-')
        data = [c.contractuel.matricule,c.contractuel.nom,c.contractuel.prenom,
                c.get_type_conge_display() if hasattr(c,'get_type_conge_display') else c.type_conge,
                c.date_debut.strftime('%d/%m/%Y') if c.date_debut else '',
                c.date_fin.strftime('%d/%m/%Y') if c.date_fin else '',
                jours,STATUS.get(c.statut,c.statut),approuve]
        fc = "F0FFF4" if c.statut=='APPROUVE' else ("FFF0F0" if c.statut=='REJETE' else "FFFBF0")
        rf = PatternFill("solid", fgColor=fc)
        for col2, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col2, value=val)
            cell.border = border; cell.fill = rf
            cell.alignment = Alignment(horizontal='center' if col2>3 else 'left', vertical='center')
    for i, ww in enumerate([14,20,20,16,14,14,8,14,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = ww
    ent_slug = entreprise.nom.replace(' ','_') if entreprise else 'toutes'
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="conges_{ent_slug}_{mois:02d}_{annee}.xlsx"'
    wb.save(resp); return resp


@login_required
def entreprise_export_permissions(request):
    """Exporte les permissions de l'entreprise en Excel.

    Args:
        request: HttpRequest Django.

    Returns:
        FileResponse : fichier Excel.
"""
    entreprise, err = ent_check(request)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installe.", status=500)
    today = date.today()
    mois  = int(request.GET.get('mois', today.month))
    annee = int(request.GET.get('annee', today.year))
    qs = Permission.objects.select_related('contractuel','approuve_par').filter(
        date_debut__month=mois, date_debut__year=annee)
    if entreprise:
        qs = qs.filter(contractuel__entreprise=entreprise)
    qs = qs.order_by('contractuel__nom')
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Permissions {mois:02d}-{annee}"
    thin = Side(style='thin', color='333344')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(bold=True, color="D4A853", size=11)
    headers = ['Matricule','Nom','Prenom','Date debut','Date fin','Motif','Statut','Approuve par']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
    ws.row_dimensions[1].height = 22
    STATUS = {'EN_ATTENTE':'En attente','APPROUVE':'Approuve','REJETE':'Rejete'}
    for row, p in enumerate(qs, 2):
        approuve = (p.approuve_par.get_full_name() or p.approuve_par.username
                    if p.approuve_par else '-')
        data = [p.contractuel.matricule,p.contractuel.nom,p.contractuel.prenom,
                p.date_debut.strftime('%d/%m/%Y') if p.date_debut else '-',
                p.date_fin.strftime('%d/%m/%Y') if p.date_fin else '-',
                p.motif or '',STATUS.get(p.statut,p.statut),approuve]
        fc = "F0FFF4" if p.statut=='APPROUVE' else ("FFF0F0" if p.statut=='REJETE' else "FFFBF0")
        rf = PatternFill("solid", fgColor=fc)
        for col2, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col2, value=val)
            cell.border = border; cell.fill = rf
            cell.alignment = Alignment(horizontal='center' if col2>3 else 'left', vertical='center')
    for i, ww in enumerate([14,20,20,14,14,28,14,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = ww
    ent_slug = entreprise.nom.replace(' ','_') if entreprise else 'toutes'
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="permissions_{ent_slug}_{mois:02d}_{annee}.xlsx"'
    wb.save(resp); return resp


@login_required
def entreprise_export_agents(request):
    """Exporte la liste des agents actifs de l'entreprise en Excel.

    Args:
        request: HttpRequest Django.

    Returns:
        FileResponse : fichier Excel.
"""
    entreprise, err = ent_check(request)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl non installe.", status=500)
    qs = Contractuel.objects.select_related('poste','direction','entreprise').filter(statut='ACTIF')
    if entreprise:
        qs = qs.filter(entreprise=entreprise)
    qs = qs.order_by('nom','prenom')
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = "Agents actifs"
    thin = Side(style='thin', color='333344')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    h_font = Font(bold=True, color="D4A853", size=11)
    alt_fill = PatternFill("solid", fgColor="F8F9FF")
    headers = ['Matricule','Nom','Prenom','Poste','Direction','Email','Telephone','Date embauche','Statut']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border
    ws.row_dimensions[1].height = 22
    for row, a in enumerate(qs, 2):
        data = [a.matricule,a.nom,a.prenom,
                a.poste.titre if a.poste else '-',
                a.direction.nom if a.direction else '-',
                a.email or '',a.telephone or '',
                a.date_embauche.strftime('%d/%m/%Y') if a.date_embauche else '','Actif']
        rf = PatternFill("solid",fgColor="FFFFFF") if row%2==0 else alt_fill
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border; cell.fill = rf
            cell.alignment = Alignment(
                horizontal='center' if col in (1,8,9) else 'left', vertical='center')
    for i, ww in enumerate([14,20,20,22,20,28,16,14,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = ww
    ent_slug = entreprise.nom.replace(' ','_') if entreprise else 'toutes'
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="agents_{ent_slug}.xlsx"'
    wb.save(resp); return resp
