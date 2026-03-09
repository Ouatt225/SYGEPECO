from django.http import HttpResponse
from .models import ActionLog, Contractuel, Presence, Conge, Permission
from .constants import (
    CONTRACTUEL_ACTIF, STATUT_APPROUVE, QUOTA_CONGE_ANNUEL,
    EXCEL_HEADER_BG, EXCEL_HEADER_FONT, EXCEL_BORDER_COLOR,
    EXCEL_ROW_APPROUVE, EXCEL_ROW_REJETE, EXCEL_ROW_ATTENTE,
    EXCEL_ROW_ALT_EVEN, EXCEL_ROW_ALT_ODD,
)



# ─── openpyxl — imports au niveau module (utilisés par les 4 exports) ────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as _gcl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_STATUS_LABELS = {
    'EN_ATTENTE':     'En attente',
    'APPROUVE':       'Approuvé',
    'REJETE':         'Rejeté',
    'VALIDE_MANAGER': 'Validé manager',
    'ANNULE':         'Annulé',
}

# ─── Protection Formula Injection (Excel / LibreOffice / Google Sheets) ─────
_FORMULA_TRIGGERS = frozenset('=+-@\t\r')


def _safe_excel_val(val):
    """
    Neutralise les injections de formules dans les cellules Excel.

    Toute valeur textuelle commençant par =, +, -, @, \t ou \r peut être
    interprétée comme une formule par Excel, LibreOffice Calc ou Google Sheets.
    On préfixe ces valeurs par une tabulation (invisible dans la cellule)
    pour forcer leur traitement en tant que texte brut.
    """
    if isinstance(val, str) and val and val[0] in _FORMULA_TRIGGERS:
        return '\t' + val
    return val

def get_manager_direction(user):
    """Retourne la Direction du manager si l'utilisateur est MANAGER, sinon None."""
    if hasattr(user, 'profile') and user.profile.role == 'MANAGER':
        return user.profile.direction
    return None


def log_action(user, action, modele='', objet_id=None, details=''):
    ActionLog.objects.create(
        utilisateur=user,
        action=action,
        modele_concerne=modele,
        objet_id=objet_id,
        details=details,
    )


def _excel_header_setup(ws, headers, col_widths):
    """Applique l'en-tête stylisé et les largeurs de colonnes. Retourne l'objet Border."""
    if not _OPENPYXL_OK:
        return None
    header_fill = PatternFill(fill_type='solid', fgColor=EXCEL_HEADER_BG)
    header_font = Font(bold=True, color=EXCEL_HEADER_FONT, size=11)
    thin = Side(style='thin', color=EXCEL_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 22
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[_gcl(i)].width = w
    return border


def _excel_response(wb, filename: str):
    """Retourne une HttpResponse en téléchargement contenant le classeur."""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _excel_row_fill(statut: str):
    """Retourne le PatternFill de la ligne selon le statut (APPROUVE / REJETE / autre)."""
    color = EXCEL_ROW_APPROUVE if statut == 'APPROUVE' else (
            EXCEL_ROW_REJETE   if statut == 'REJETE'   else EXCEL_ROW_ATTENTE)
    return PatternFill("solid", fgColor=color)


def solde_conges_annuel(contractuel, annee=None):
    """Calcule le solde de conges restant pour un contractuel."""
    from datetime import date as _date
    if annee is None:
        annee = _date.today().year
    jours_pris = sum(
        cg.nb_jours() for cg in contractuel.conges.filter(
            type_conge='ANNUEL', statut=STATUT_APPROUVE, date_debut__year=annee,
        ) if cg.date_debut and cg.date_fin
    )
    return max(0, QUOTA_CONGE_ANNUEL - jours_pris)


def export_presences_excel(mois, annee, direction=None):
    if not _OPENPYXL_OK:
        return HttpResponse("openpyxl non installé.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presences {mois:02d}-{annee}"

    col_widths = [14, 20, 20, 20, 12, 12, 12, 12]
    headers    = ['Matricule', 'Nom', 'Prénom', 'Direction', 'Présents', 'Absents', 'Retards', 'Taux (%)']
    border     = _excel_header_setup(ws, headers, col_widths)

    contractuels = Contractuel.objects.filter(statut='ACTIF').select_related('direction')
    if direction:
        contractuels = contractuels.filter(direction=direction)

    for row, c in enumerate(contractuels, 2):
        presences  = Presence.objects.filter(contractuel=c, date__month=mois, date__year=annee)
        nb_present = presences.filter(statut='PRESENT').count()
        nb_absent  = presences.filter(statut='ABSENT').count()
        nb_retard  = presences.filter(statut='RETARD').count()
        total      = nb_present + nb_absent + nb_retard
        taux       = round(nb_present / total * 100, 1) if total > 0 else 0

        data = [c.matricule, c.nom, c.prenom,
                c.direction.nom if c.direction else '—',
                nb_present, nb_absent, nb_retard, taux]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=_safe_excel_val(val))
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col > 3 else 'left')

    return _excel_response(wb, f"presences_{mois:02d}_{annee}.xlsx")


def export_conges_excel(mois=None, annee=None, direction=None):
    if not _OPENPYXL_OK:
        return HttpResponse("openpyxl non installé.", status=500)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"Conges {f'{mois:02d}-{annee}' if mois and annee else 'Tous'}"

    col_widths = [14, 20, 20, 22, 16, 14, 14, 8, 14, 22]
    headers    = ['Matricule', 'Nom', 'Prénom', 'Entreprise', 'Type',
                  'Date début', 'Date fin', 'Jours', 'Statut', 'Approuvé par']
    border     = _excel_header_setup(ws, headers, col_widths)

    qs = Conge.objects.select_related('contractuel', 'contractuel__entreprise', 'approuve_par')
    if mois and annee:
        qs = qs.filter(date_debut__month=mois, date_debut__year=annee)
    if direction:
        qs = qs.filter(contractuel__direction=direction)
    qs = qs.order_by('contractuel__nom')

    for row, c in enumerate(qs, 2):
        nb_jours = (c.date_fin - c.date_debut).days + 1 if c.date_fin and c.date_debut else ''
        approuve = c.approuve_par.get_full_name() or c.approuve_par.username if c.approuve_par else '—'
        data = [
            c.contractuel.matricule,
            c.contractuel.nom,
            c.contractuel.prenom,
            c.contractuel.entreprise.nom if c.contractuel.entreprise else '—',
            c.get_type_conge_display() if hasattr(c, 'get_type_conge_display') else c.type_conge,
            c.date_debut.strftime('%d/%m/%Y') if c.date_debut else '',
            c.date_fin.strftime('%d/%m/%Y')   if c.date_fin   else '',
            nb_jours,
            _STATUS_LABELS.get(c.statut, c.statut),
            approuve,
        ]
        row_fill = _excel_row_fill(c.statut)
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=_safe_excel_val(val))
            cell.border = border
            cell.fill   = row_fill
            cell.alignment = Alignment(horizontal='center' if col > 3 else 'left', vertical='center')

    fname = f"conges_{mois:02d}_{annee}.xlsx" if mois and annee else "conges_tous.xlsx"
    return _excel_response(wb, fname)


def export_permissions_excel(mois=None, annee=None, direction=None):
    if not _OPENPYXL_OK:
        return HttpResponse("openpyxl non installé.", status=500)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = f"Permissions {f'{mois:02d}-{annee}' if mois and annee else 'Tous'}"

    col_widths = [14, 20, 20, 22, 14, 14, 30, 14, 22]
    headers    = ['Matricule', 'Nom', 'Prénom', 'Entreprise',
                  'Date début', 'Date fin', 'Motif', 'Statut', 'Approuvé par']
    border     = _excel_header_setup(ws, headers, col_widths)

    qs = Permission.objects.select_related('contractuel', 'contractuel__entreprise', 'approuve_par')
    if mois and annee:
        qs = qs.filter(date_debut__month=mois, date_debut__year=annee)
    if direction:
        qs = qs.filter(contractuel__direction=direction)
    qs = qs.order_by('contractuel__nom')

    for row, p in enumerate(qs, 2):
        approuve = p.approuve_par.get_full_name() or p.approuve_par.username if p.approuve_par else '—'
        data = [
            p.contractuel.matricule,
            p.contractuel.nom,
            p.contractuel.prenom,
            p.contractuel.entreprise.nom if p.contractuel.entreprise else '—',
            p.date_debut.strftime('%d/%m/%Y') if p.date_debut else '',
            p.date_fin.strftime('%d/%m/%Y')   if p.date_fin   else '',
            p.motif or '—',
            _STATUS_LABELS.get(p.statut, p.statut),
            approuve,
        ]
        row_fill = _excel_row_fill(p.statut)
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=_safe_excel_val(val))
            cell.border = border
            cell.fill   = row_fill
            cell.alignment = Alignment(horizontal='center' if col > 3 else 'left', vertical='center')

    fname = f"permissions_{mois:02d}_{annee}.xlsx" if mois and annee else "permissions_tous.xlsx"
    return _excel_response(wb, fname)


def export_contractuels_excel(direction=None):
    if not _OPENPYXL_OK:
        return HttpResponse("openpyxl non installé.", status=500)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Contractuels actifs"

    col_widths = [14, 20, 20, 10, 22, 20, 20, 28, 16, 14, 10]
    headers    = ['Matricule', 'Nom', 'Prénom', 'Genre', 'Entreprise',
                  'Direction', 'Poste', 'Email', 'Téléphone', 'Date embauche', 'Statut']
    border     = _excel_header_setup(ws, headers, col_widths)

    qs = Contractuel.objects.filter(statut='ACTIF').select_related(
        'entreprise', 'direction', 'poste'
    ).order_by('entreprise__nom', 'nom')
    if direction:
        qs = qs.filter(direction=direction)

    for row, c in enumerate(qs, 2):
        fill_color = EXCEL_ROW_ALT_EVEN if row % 2 == 0 else EXCEL_ROW_ALT_ODD
        row_fill   = PatternFill("solid", fgColor=fill_color)
        data = [
            c.matricule,
            c.nom,
            c.prenom,
            'Homme' if c.genre == 'M' else 'Femme',
            c.entreprise.nom if c.entreprise else '—',
            c.direction.nom  if c.direction  else '—',
            c.poste.titre    if c.poste      else '—',
            c.email     or '—',
            c.telephone or '—',
            c.date_embauche.strftime('%d/%m/%Y') if c.date_embauche else '',
            'Actif',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=_safe_excel_val(val))
            cell.border = border
            cell.fill   = row_fill
            cell.alignment = Alignment(horizontal='center' if col in (4, 10, 11) else 'left', vertical='center')

    return _excel_response(wb, "contractuels_actifs.xlsx")


def build_fiche_pdf(contractuel):
    """Construit la fiche PDF d'un agent. Retourne un BytesIO pret a lire."""
    import os as _os
    from io import BytesIO
    from datetime import date as _date
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image)

    c = contractuel
    today = _date.today()
    GREEN    = colors.HexColor('#3E7B52')
    GREEN_TXT = colors.HexColor('#2D6B42')
    WHITE    = colors.white
    BLACK    = colors.HexColor('#1A1A1A')
    GREY_LT  = colors.HexColor('#F7F7F7')
    GREY_MID = colors.HexColor('#CCCCCC')

    def S(name, **kw): return ParagraphStyle(name, **kw)
    lbl = S('lbl', fontName='Helvetica-Bold', fontSize=8.5, textColor=BLACK, leading=11)
    val = S('val', fontName='Helvetica',      fontSize=8.5, textColor=GREEN_TXT, leading=11)
    sec = S('sec', fontName='Helvetica-Bold', fontSize=10,  textColor=WHITE, leading=14)
    mat = S('mat', fontName='Helvetica-Bold', fontSize=9,   textColor=BLACK, alignment=TA_CENTER)
    ini = S('ini', fontName='Helvetica-Bold', fontSize=24,  textColor=GREEN,
            alignment=TA_CENTER, leading=30)

    W_total = A4[0] - 2.4*cm
    W_left  = 13.1*cm
    W_right =  5.5*cm
    CW4 = [3.6*cm, 5.5*cm, 3.6*cm, 5.9*cm]

    def sec_hdr(title):
        t = Table([[Paragraph(title, sec)]], colWidths=[W_total])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), GREEN),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
        ]))
        return t

    def grid4(rows, span_rows=None):
        data = []; spans = []
        for i, r in enumerate(rows):
            if span_rows and i in span_rows:
                data.append([Paragraph(r[0], lbl), Paragraph(str(r[1]) if r[1] else '-', val), '', ''])
                spans.append(('SPAN', (1, i), (3, i)))
            else:
                data.append([
                    Paragraph(r[0], lbl), Paragraph(str(r[1]) if r[1] else '-', val),
                    Paragraph(r[2] if len(r) > 2 else '', lbl),
                    Paragraph(str(r[3]) if len(r) > 3 and r[3] else '-', val),
                ])
        t = Table(data, colWidths=CW4)
        t.setStyle(TableStyle([
            ('TOPPADDING',    (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 8), ('RIGHTPADDING',  (0,0), (-1,-1), 4),
            ('LINEBELOW',  (0,0), (-1,-2), 0.4, GREY_MID),
            ('LINEBEFORE', (2,0), (2,-1),  0.4, GREY_MID),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [WHITE, GREY_LT]),
            ('BOX',   (0,0), (-1,-1), 0.5, GREY_MID),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ] + spans))
        return t

    # --- Données calculées ---
    age = annee_retraite = date_retraite = None
    if c.date_naissance:
        b = c.date_naissance
        age = today.year - b.year - ((today.month, today.day) < (b.month, b.day))
        annee_retraite = b.year + 60
        try:    date_retraite = _date(annee_retraite, b.month, b.day)
        except ValueError: date_retraite = _date(annee_retraite, b.month, 28)

    conges_pris = sum(
        (cg.date_fin - cg.date_debut).days + 1
        for cg in c.conges.filter(statut='APPROUVE')
        if cg.date_debut and cg.date_fin)
    solde_conges  = max(0, 30 - conges_pris)
    contrat       = c.get_contrat_actif()
    sal           = int(contrat.salaire) if contrat and contrat.salaire else None
    salaire_str   = f"{sal:,} FCFA".replace(",", " ") if sal else '-'
    entreprise_nom = c.entreprise.nom if getattr(c, 'entreprise', None) else '-'
    commune       = getattr(c, 'commune',      None) or '-'
    ville         = getattr(c, 'ville',        None) or '-'
    cnps          = getattr(c, 'numero_cnps',  None) or '-'
    nb_enfants    = str(c.nombre_enfants) if hasattr(c, 'nombre_enfants') else '-'
    try:    sit_fam = c.get_situation_famille_display()
    except Exception: sit_fam = '-'

    presence_today = c.presences.filter(date=today).first()
    if presence_today:      etat_str = presence_today.get_statut_display().upper()
    elif c.statut == 'ACTIF': etat_str = 'PRESENT ET PAYE'
    else:                      etat_str = c.get_statut_display().upper()

    # --- Construction du document ---
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=1.2*cm, rightMargin=1.2*cm,
        topMargin=1.2*cm,  bottomMargin=1.2*cm,
        title=f'Fiche {c.nom} {c.prenom}')
    story = []

    story.append(sec_hdr('INFORMATIONS GENERALES'))
    left_data = [[Paragraph(r[0], lbl), Paragraph(str(r[1]), val)] for r in [
        ('Matricule :',        c.matricule),
        ('Nom et Prenoms :',   f'{c.nom} {c.prenom}'),
        ('Date de Naissance :', c.date_naissance.strftime('%d/%m/%Y') if c.date_naissance else '-'),
        ('Lieu de Naissance :', c.lieu_naissance or '-'),
        ('Age actuel :',        f'{age} ANS' if age is not None else '-'),
        ('Annee de Retraite :', str(annee_retraite) if annee_retraite else '-'),
    ]]
    left_tbl = Table(left_data, colWidths=[3.8*cm, 9.3*cm])
    left_tbl.setStyle(TableStyle([
        ('TOPPADDING',    (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 8), ('RIGHTPADDING',  (0,0), (-1,-1), 4),
        ('LINEBELOW',     (0,0), (-1,-2), 0.4, GREY_MID),
        ('ROWBACKGROUNDS',(0,0), (-1,-1), [WHITE, GREY_LT]),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))

    # Protection path traversal : vérifier que la photo est sous MEDIA_ROOT
    _photo_ok = False
    if c.photo and hasattr(c.photo, 'path'):
        try:
            from pathlib import Path
            from django.conf import settings as _settings
            _resolved   = Path(c.photo.path).resolve()
            _media_root = Path(_settings.MEDIA_ROOT).resolve()
            _photo_ok = _resolved.is_relative_to(_media_root) and _resolved.exists()
        except (ValueError, OSError):
            _photo_ok = False
    if _photo_ok:
        photo_el = Image(str(Path(c.photo.path).resolve()), width=3.8*cm, height=4.2*cm)
    else:
        photo_el = Paragraph(f'{c.nom[:1]}{c.prenom[:1]}'.upper(), ini)

    right_tbl = Table([[photo_el], [Paragraph(c.matricule, mat)]], colWidths=[W_right])
    right_tbl.setStyle(TableStyle([
        ('ALIGN',   (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (0,0), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LINEBELOW',     (0,0), (0,0),  0.4, GREY_MID),
        ('ROWBACKGROUNDS',(0,0), (-1,-1), [WHITE, GREY_LT]),
    ]))

    hero = Table([[left_tbl, right_tbl]], colWidths=[W_left, W_right])
    hero.setStyle(TableStyle([
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0), ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 0), ('BOTTOMPADDING',(0,0), (-1,-1), 0),
        ('BOX',      (0,0), (-1,-1), 0.5, GREY_MID),
        ('LINEBEFORE',(1,0), (1,-1),  0.5, GREY_MID),
    ]))
    story.append(hero)

    story.append(grid4([
        ('Sexe :',         c.get_genre_display(), 'Email :',         c.email or '-'),
        ('Numero CNPS :',  cnps,                  'Commune :',       commune),
        ('Telephone :',    c.telephone or '-',   "Nombre d'enfant :", nb_enfants),
        ('Ville :',        ville,                 '',                 ''),
        ('Situation Famille :', sit_fam,           '',                 ''),
    ]))
    story.append(Spacer(1, 0.25*cm))

    story.append(sec_hdr('EMPLOI'))
    dir_nom     = c.direction.nom if c.direction else '-'
    lieu_travail = ville if ville != '-' else (getattr(c, 'adresse', None) or '-')
    story.append(grid4([
        ('Entreprise :', entreprise_nom, "Date d'embauche :",
         c.date_embauche.strftime('%d/%m/%Y') if c.date_embauche else '-'),
        ('Direction :',  dir_nom,        '', ''),
        ('Salaire :',    salaire_str,    'Lieu de Travail :', lieu_travail),
        ('Emploi :',     c.poste.titre if c.poste else '-', '', ''),
    ], span_rows={1, 3}))
    story.append(Spacer(1, 0.25*cm))

    story.append(sec_hdr('ETAT AGENT'))
    story.append(grid4([
        ('Date prise de service :',
         c.date_embauche.strftime('%d/%m/%Y') if c.date_embauche else '-',
         'Solde Conges :', f'{solde_conges} JOURS'),
        ('Date depart retraite :',
         date_retraite.strftime('%d/%m/%Y') if date_retraite else '-',
         'Conges Pris :', f'{conges_pris} JOURS'),
        ('Etat :', etat_str, '', ''),
    ], span_rows={2}))

    doc.build(story)
    buf.seek(0)
    return buf
